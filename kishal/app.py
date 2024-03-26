from flask import Flask, render_template, request
import openpyxl

app = Flask(__name__)

# Create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Define column headers
worksheet['A1'] = 'Name'
worksheet['B1'] = 'Phone'
worksheet['C1'] = 'Email'
worksheet['D1'] = 'Event'

# Set the starting row for data
row = 2

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/register', methods=['POST'])
def register():
    global row

    # Get form data
    name = request.form['name']
    phone = request.form['phone']
    email = request.form['email']
    event = request.form['event']

    # Write data to the worksheet
    worksheet.cell(row=row, column=1, value=name)
    worksheet.cell(row=row, column=2, value=phone)
    worksheet.cell(row=row, column=3, value=email)
    worksheet.cell(row=row, column=4, value=event)

    # Save the workbook
    workbook.save('registration_data.xlsx')

    # Increment row for next entry
    row += 1

    # Render the success page
    return render_template('success.html')

if __name__ == '__main__':
    app.run(debug=True)